import pdfplumber
from utils import *
from gui import *
import re
import os
import openpyxl
import yaml


# --------------------------------------------------------------------------------
#                                    PDF
# --------------------------------------------------------------------------------
class PDF:
    def __init__(self, pdf_file_name, keywords_, summary_pages_ajustment):
        self.pdf = pdfplumber.open(pdf_file_name)
        self.keywords_to_search_on_summary = keywords_
        self.summary_pages_ajustment = summary_pages_ajustment  # ajustment of real page number and page indicated on the summary
        self.last_page_to_look_for_summary = 25


    def search_for_summary(self):
        summary_page = 0
        for page in range(0, self.last_page_to_look_for_summary):
            text = self.pdf.pages[page].extract_text()
            if ('sumário' in text.lower()) and '.....' in text:
                summary_page = page
                break
        return summary_page


    def create_dict_from_summary(self):
        all_text = ''
        summary_page = self.search_for_summary()
        for num in range(summary_page, summary_page + 10):
            page = self.pdf.pages[num]
            width = page.width
            height = page.height
            reading_coordinate = (0, margin_top, width, height - margin_bottom)

            text = page.within_bbox(reading_coordinate).extract_text()

            for identifier in summary_identifiers:
                if identifier in text:
                    all_text += text
                    all_text += '\n'
        #print(all_text)

        #summary_number_regex = r'\b\d+\.\d*\s'
        title_page_regex = r'(.+?)\s*\.{3,}\s*(\d+)'
        title_page_underline_regex = r'(.+?)\s*\_{3,}\s*(\d+)'

        #all_text = re.sub(summary_number_regex, '', all_text)

        title_page = re.findall(title_page_regex, all_text)
        if not title_page:
            title_page = re.findall(title_page_underline_regex, all_text)

        #for title, page in title_page:
            #print("Título:", title)
            #print("Página:", page)
            #print()

        return title_page


    def search_on_summary_titles(self, summary_titles):
        titles_dict = summary_titles
        pages_list = []
        #print("\nPalavras a serem buscadas no sumário: ", self.keywords_to_search_on_summary)
        for keyword in self.keywords_to_search_on_summary:
            for index, (title, page) in enumerate(titles_dict):
                if keyword in title.lower():
                    #print(f"A palavra-chave '{keyword}' foi encontrada no título '{title}' na página {page} na chave {index}\n\n")
                    pages_list.append(str(page) + '-' + str(list(titles_dict)[index + 2][1]))
                    del titles_dict[index]
        return pages_list


    def exctract_text_from_pdf(self, pages_to_extract_text):
        all_text = ''
        for pages_interval in pages_to_extract_text:
            first_page, last_page = pages_interval.split('-')
            first_page = int(first_page) - 1 + self.summary_pages_ajustment
            last_page = int(last_page) + self.summary_pages_ajustment
            for num in range(first_page, last_page):
                page = self.pdf.pages[num]
                width = page.width
                height = page.height
                reading_coordinate = (0, margin_top, width, height - margin_bottom)

                #print(f"\n\n--------------------- PÁGINA {num + 1} ---------------------\n")
                if self.has_table(page):
                    #print("****** POSSUI TABELA ******\n")
                    tables = page.extract_tables()
                    tables_correct_text = []
                    tables_extracted_text = []
                    all_text_with_tables = page.within_bbox(reading_coordinate).extract_text()

                    for table in tables:
                        tables_correct_text.append(self.extract_tables_text(table))

                    tables_box = page.find_tables()
                    
                    for table_box in tables_box:
                        box = table_box.bbox
                        tables_extracted_text.append(page.within_bbox(reading_coordinate).within_bbox(box).extract_text())
                        
                    text = replace_table_for_text(all_text_with_tables, tables_extracted_text, tables_correct_text)
                else:
                    text = page.within_bbox(reading_coordinate).extract_text().replace('\n', ' ')

                all_text += text
                #print(text)

        all_text = clean_text(all_text)
        #print('\n\n\n\n', '-------------------- TEXTO FINAL --------------------\n\n\n', all_text)
        return all_text


    def has_table(self, page):
        tables = page.extract_tables()
        if tables:
            return True
        return False


    def extract_tables_text(self, table):
        table_text = 'Tabela: '
        for row in table:
            row_reading = ['' if item is None else item for item in row]
            row_reading = [item.replace('\n', ' ') for item in row_reading]

            result = str(row_reading)
            table_text += result
            table_text += '\\n'

        return table_text


    def write_reduced_text(self, text, file_name='reduced_text'):
        file_name = file_name + '.txt'
        reduced_text_path = os.path.join(os.getcwd(), file_name)
        with open(reduced_text_path, 'w') as file:
            file.write(text)
        return

# --------------------------------------------------------------------------------
#                                    Sheet
# --------------------------------------------------------------------------------
class Sheet:
    def __init__(self, sheet_path):
        self.sheet_path = sheet_path + '/Prognóstico_FUNASA.xlsx'
        self.yaml_path = os.path.join(os.getcwd(), 'sheet_config.yaml')
        self.wb = openpyxl.Workbook()
        self.config = self.get_config_from_yaml()


    def create_sheet(self):
        #TODO: treat if exists
        if not os.path.exists(self.sheet_path):
            self.wb.save(self.sheet_path)
        self.fill_sheet_before_data()            
        return
    

    def get_config_from_yaml(self):
        with open(self.yaml_path, 'r') as file:
            config = yaml.safe_load(file)
        return config


    def fill_sheet_before_data(self):
        sheet_names = self.wb.sheetnames
        for sheet in self.config:
            if not sheet in sheet_names:
                ws = self.wb.create_sheet(sheet)
            else:
                ws = self.wb[sheet]
            for cell, values in self.config[sheet].items():
                ws[cell] = values['value']
                self.apply_cell_style(ws[cell], values['style'])
        if 'Sheet' in sheet_names:
            self.wb.remove(self.wb['Sheet'])
        self.wb.save(self.sheet_path)
        return


    def apply_cell_style(self, cell, style):
        if not style:
            return
        if 'bold' in style:
            cell.font = openpyxl.styles.Font(bold=True)
        if 'italic' in style:
            cell.font = openpyxl.styles.Font(italic=True)
        return


    def search_on_text(self, sheet_name, text):
        for cell, values in self.config[sheet_name].items():
            if not 'keys' in values:
                continue
            keywords = values['keys']
            approched = False
            for keyword in keywords:
                if keyword in text:
                    #print("\nkeyword: ", keyword)
                    approched = True
                    break

            self.write_on_sheet(sheet_name, cell, approched)

        return


    def write_on_sheet(self, sheet_name, cell, approched):
        next_cell = chr(ord(cell[0]) + 1) + cell[1:]
        ws = self.wb[sheet_name]
        ws[next_cell] = 'Abordado' if approched else 'Não Abordado'
        self.wb.save(self.sheet_path)
        return


# --------------------------------------------------------------------------------
#                                    main
# --------------------------------------------------------------------------------
if __name__ == '__main__':
    gui = GUI()
    config = gui.gui()
    #print(config)
    pdf_file_name = config.get('pdf_file_name')
    funasa_dict = config.get('funasa_dict')
    keywords_ = treat_if_is_empty(config.get('keywords_obj'), keywords_default)
    summary_pages_ajustment = treat_if_is_empty(config.get('summary_pages_ajustment'), 0)

    if not pdf_file_name:
        erro_popup('Plano não selecionado')
        exit(1)
    if not funasa_dict:
        erro_popup('Diretório não selecionado')
        exit(1)

    print("Gerando texto simplificando para a inteligência...")
    pdf = PDF(pdf_file_name, keywords_, int(summary_pages_ajustment))
    summary_titles = pdf.create_dict_from_summary()
    pages_to_extract_text = pdf.search_on_summary_titles(summary_titles)
    text_for_intel = pdf.exctract_text_from_pdf(pages_to_extract_text)
    pdf.write_reduced_text(text_for_intel)

    print("Criando planilha FUNASA...")
    sheet = Sheet(funasa_dict)
    sheet.create_sheet()

    print("Analisando temas abordados...")
    for sheet_name, value in keywords_approched_or_not.items():
        pdf_approched = PDF(pdf_file_name, keywords_approched_or_not[sheet_name], int(summary_pages_ajustment))
        summary_titles_approched = pdf_approched.create_dict_from_summary()
        pages_to_extract_text_approched = pdf_approched.search_on_summary_titles(summary_titles_approched)
        text_approched = pdf_approched.exctract_text_from_pdf(pages_to_extract_text_approched)
        #pdf_approched.write_reduced_text(text_approched, (sheet_name + '_text_approched'))
        sheet.search_on_text(sheet_name, (text_approched).lower())